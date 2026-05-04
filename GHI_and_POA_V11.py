import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from pathlib import Path
import sys
import shutil
from datetime import datetime

# pvlib required for precise solar noon
try:
    import pvlib
except ImportError:
    pvlib = None
    print("Warning: pvlib not installed. Solar noon calculations will fail.")


# ------------------------------------------------------------
# 1) CONFIGURATION
# ------------------------------------------------------------
input_folder = Path("inputs/GHI&GHI_tilt")

# Comparison Dates (day1 vs day2)
CompareDates = ("2026-04-30", "2026-05-01")

# --- DYNAMIC OUTPUT FOLDER CREATION ---
script_name = Path(__file__).stem
# Define the latest date (day2) as the run identifier
run_identifier_date = CompareDates[1] 
# Construct the path: outputs / ScriptName / LatestDate
output_folder = Path("outputs") / script_name / run_identifier_date
output_folder.mkdir(parents=True, exist_ok=True)


# Time Window for Plots
start_time_limit = "7:00:00"
end_time_limit = "18:00:00"

# Site Coords / Timezone
LAT = 40.26
LON = -83.99
TZ = "Etc/GMT+5"  # Locked to Standard Time (UTC-5)

# Data Smoothing (Assumes 1-minute logging frequency)
AVERAGING_WINDOW_MINS = 1

# Export Options
EXPORT_ESTIMATED_NOONS = True

# Station Filtering & Styling
SENSORS_TO_INCLUDE = [
    # POA Sensors
  "MET02/POA_1", "MET02/POA_2",
   "MET16/POA_1", "MET16/POA_2",
    "MET22/POA_1", "MET22/POA_2",
   "MET37/POA_1", "MET37/POA_2",
    
    # RPOA Sensors
   "MET02/RPOA_1", "MET02/RPOA_2",
   "MET16/RPOA_1", "MET16/RPOA_2",
   "MET22/RPOA_1", "MET22/RPOA_2",
   "MET37/RPOA_1", "MET37/RPOA_2",

    # GHI Sensors
    "MET22/GHI",
    "MET02/GHI",
    "MET16/GHI",
    "MET37/GHI"
]

station_colors = {
    "MET02": "tab:blue",
    "MET16": "tab:green",
    "MET22": "tab:red",
    "MET37": "tab:orange",
}
line_styles = {"POA_1": "-", "POA_2": ":", "GHI": "-", "RPOA_1": "-", "RPOA_2": ":"}

# Tick Formatting
MAJOR_TICK_MINUTES = 15
MINOR_TICK_MINUTES = 5


# ------------------------------------------------------------
# 2) HELPER FUNCTIONS
# ------------------------------------------------------------

def parse_date(date_str: str):
    return datetime.strptime(date_str, "%Y-%m-%d").date()

def prep_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["t_stamp_dt"] = pd.to_datetime(df["t_stamp"], errors="coerce")
    df = df.sort_values("t_stamp_dt").reset_index(drop=True)
    
    if AVERAGING_WINDOW_MINS > 1:
        num_cols = df.select_dtypes(include='number').columns
        df[num_cols] = df[num_cols].rolling(window=AVERAGING_WINDOW_MINS, center=True, min_periods=1).mean()
        
    df["date"] = df["t_stamp_dt"].dt.date
    return df

def filter_day_time(df: pd.DataFrame, day):
    day_df = df[df["date"] == day].copy()
    start_ts = pd.to_datetime(f"{day} {start_time_limit}")
    end_ts = pd.to_datetime(f"{day} {end_time_limit}")
    return day_df[(day_df["t_stamp_dt"] >= start_ts) & (day_df["t_stamp_dt"] <= end_ts)].copy()

def get_precise_solar_noon(date_obj):
    if not pvlib:
        return datetime.combine(date_obj, pd.to_datetime("12:00:00").time())
    
    times = pd.date_range(start=f"{date_obj} 11:00:00", end=f"{date_obj} 15:00:00", freq="1min", tz=TZ)
    solpos = pvlib.solarposition.get_solarposition(times, LAT, LON)
    noon_idx = solpos['elevation'].idxmax()
    return noon_idx.to_pydatetime()

def format_time_axis(ax):
    ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=MAJOR_TICK_MINUTES))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_minor_locator(mdates.MinuteLocator(interval=MINOR_TICK_MINUTES))
    ax.grid(True, which="major", linestyle="--", alpha=0.35)
    ax.grid(True, which="minor", linestyle="--", alpha=0.15)

def add_stats_box(ax, target_row, cols, label, unit):
    if target_row is None or target_row.empty or not cols:
        return

    present_cols = [c for c in cols if c in target_row.columns]
    if not present_cols:
        return

    vals = target_row[present_cols].iloc[0].astype(float)
    v_min, v_max = vals.min(), vals.max()
    delta = v_max - v_min

    stats_text = (
        f"{label} @ Noon\n"
        f"Median: {vals.median():.2f}{unit}\n"
        f"Average: {vals.mean():.2f}{unit}\n"
        f"Range: {v_min:.1f} - {v_max:.1f}{unit}\n"
        f"Delta (Δ): {delta:.1f}{unit}"
    )

    if "W/m" in unit:
        if v_min != 0:
            pct_dev = (delta / v_min) * 100
            check_str = "YES" if pct_dev < 2.0 else "NO"
            stats_text += f"\nΔ % of Min: {pct_dev:.2f}%"
            stats_text += f"\n< 2% Check: {check_str}"
        else:
            stats_text += f"\nΔ % of Min: N/A"
            stats_text += f"\n< 2% Check: N/A"

    ax.text(
        0.01, 0.98, stats_text,
        transform=ax.transAxes,
        fontsize=9,
        verticalalignment="top",
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.85)
    )

def add_albedo_box(ax, target_row, stations):
    """
    Calculates Effective Albedo (Avg RPOA / Avg POA) at solar noon for each station.
    Strictly filters out <= 0 and NaN values prior to averaging.
    Calculates and displays the median across all valid stations.
    """
    if target_row is None or target_row.empty:
        return

    text_lines = ["Effective Albedo @ Noon"]
    valid_albedos = []

    for st in stations:
        rpoa_cols = [c for c in target_row.columns if c.startswith(f"{st}/RPOA") and "TILT" not in c]
        poa_cols = [c for c in target_row.columns if c.startswith(f"{st}/POA") and "TILT" not in c]

        if rpoa_cols and poa_cols:
            r_vals = target_row[rpoa_cols].iloc[0].astype(float)
            p_vals = target_row[poa_cols].iloc[0].astype(float)

            # Ignore 0 or negative values. Drop nulls.
            r_valid = r_vals[r_vals > 0].dropna()
            p_valid = p_vals[p_vals > 0].dropna()

            if not r_valid.empty and not p_valid.empty:
                # If only one valid sensor remains, mean() safely uses that single value
                avg_rpoa = r_valid.mean()
                avg_poa = p_valid.mean()
                albedo_pct = (avg_rpoa / avg_poa) * 100
                text_lines.append(f"{st}: {albedo_pct:.1f}%")
                valid_albedos.append(albedo_pct)
            else:
                text_lines.append(f"{st}: N/A")

    # Add robust median calculation if at least one station is valid
    if valid_albedos:
        median_albedo = pd.Series(valid_albedos).median()
        text_lines.append("----------------")
        text_lines.append(f"Median: {median_albedo:.1f}%")

    if len(text_lines) > 1:
        stats_text = "\n".join(text_lines)
        ax.text(
            0.22, 0.98, stats_text,
            transform=ax.transAxes,
            fontsize=9,
            verticalalignment="top",
            bbox=dict(boxstyle="round", facecolor="white", alpha=0.85)
        )

def set_top_right_legend(ax, title="Sensors"):
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(
            loc="upper left", bbox_to_anchor=(1.02, 1), 
            title=title, borderaxespad=0., fontsize=9
        )


# ------------------------------------------------------------
# 3) DATA PROCESSING & PLOTTING FUNCTIONS
# ------------------------------------------------------------

def find_best_source_for_day(loaded_items, target_day):
    for item in loaded_items:
        df_filtered = filter_day_time(item["df"], target_day)
        if not df_filtered.empty:
            return {"df_filtered": df_filtered, **item}
    return None

def calculate_medians(df, ghi_cols, poa_cols):
    df_med = df[["t_stamp_dt", "t_stamp"]].copy()
    avail_ghi = [c for c in ghi_cols if c in df.columns]
    avail_poa = [c for c in poa_cols if c in df.columns]
    
    if avail_ghi: df_med["Median_GHI"] = df[avail_ghi].median(axis=1)
    if avail_poa: df_med["Median_POA"] = df[avail_poa].median(axis=1)
        
    return df_med

def plot_compare_2x2(df1, df2, cols1, tilts1, cols2, tilts2,
                     day1, day2, noon1, noon2, noon_row1, noon_row2,
                     title_prefix, filename_suffix, mode="POA"):
    fig, axes = plt.subplots(2, 2, figsize=(20, 10), sharey="row", sharex="col")
    ax_irr_1, ax_irr_2 = axes[0, 0], axes[0, 1]
    ax_tilt_1, ax_tilt_2 = axes[1, 0], axes[1, 1]
    
    noon1_naive = noon1.replace(tzinfo=None)
    noon2_naive = noon2.replace(tzinfo=None)
    noon1_str = pd.to_datetime(noon1).round("1min").strftime("%H:%M")
    noon2_str = pd.to_datetime(noon2).round("1min").strftime("%H:%M")
    
    # --- Day 1 (Left) ---
    x1 = df1["t_stamp_dt"]
    for c in cols1:
        if c in df1.columns:
            st = c.split("/")[0]
            ls = "-"
            if "POA_2" in c or "RPOA_2" in c: ls = ":"
            label_name = c.split("/")[1] if mode in ("POA", "RPOA") else st
            ax_irr_1.plot(x1, df1[c], label=f"{st} {label_name}", 
                          color=station_colors.get(st, "black"), linestyle=ls)
            
    ax_irr_1.axvline(noon1_naive, color="red", ls="--")
    ax_irr_1.set_title(f"{mode} • {day1} • Noon: {noon1_str}", fontsize=12)
    ax_irr_1.set_ylabel(f"{mode} [W/m²]")
    add_stats_box(ax_irr_1, noon_row1, cols1, mode, " W/m²")

    if mode == "RPOA":
        stations1 = sorted(list(set([c.split("/")[0] for c in cols1])))
        add_albedo_box(ax_irr_1, noon_row1, stations1)

    for c in tilts1:
        if c in df1.columns:
            st = c.split("/")[0]
            label_name = "Tilt"
            ax_tilt_1.plot(x1, df1[c], label=f"{st} {label_name}",
                           color=station_colors.get(st, "black"))
    ax_tilt_1.axvline(noon1_naive, color="red", ls="--")
    ax_tilt_1.set_ylabel("Tilt [°]")
    add_stats_box(ax_tilt_1, noon_row1, tilts1, "Tilt", "°")

    # --- Day 2 (Right) ---
    x2 = df2["t_stamp_dt"]
    for c in cols2:
        if c in df2.columns:
            st = c.split("/")[0]
            ls = "-"
            if "POA_2" in c or "RPOA_2" in c: ls = ":"
            label_name = c.split("/")[1] if mode in ("POA", "RPOA") else st
            ax_irr_2.plot(x2, df2[c], label=f"{st} {label_name}",
                          color=station_colors.get(st, "black"), linestyle=ls)

    ax_irr_2.axvline(noon2_naive, color="red", ls="--")
    ax_irr_2.set_title(f"{mode} • {day2} • Noon: {noon2_str}", fontsize=12)
    add_stats_box(ax_irr_2, noon_row2, cols2, mode, " W/m²")

    if mode == "RPOA":
        stations2 = sorted(list(set([c.split("/")[0] for c in cols2])))
        add_albedo_box(ax_irr_2, noon_row2, stations2)

    for c in tilts2:
        if c in df2.columns:
            st = c.split("/")[0]
            ax_tilt_2.plot(x2, df2[c], label=f"{st} {label_name}",
                           color=station_colors.get(st, "black"))
    ax_tilt_2.axvline(noon2_naive, color="red", ls="--")
    add_stats_box(ax_tilt_2, noon_row2, tilts2, "Tilt", "°")

    # Formatting
    for ax in axes.flatten():
        format_time_axis(ax)
        if ax in [ax_tilt_1, ax_tilt_2]:
            ax.set_xlabel("Time (HH:MM)")
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right")
        else:
            plt.setp(ax.get_xticklabels(), visible=False)
        set_top_right_legend(ax)

    fig.suptitle(title_prefix, fontsize=13, y=0.98)
    plt.tight_layout(rect=[0, 0, 1, 0.92])
    
    save_path = output_folder / f"{mode}_COMPARE_{day1}_VS_{day2}_{filename_suffix}.png"
    plt.savefig(save_path, dpi=150)
    plt.close(fig)
    print(f"Saved: {save_path.name}")

def plot_median_ghi_poa_overlay(df, day, noon, start_time, end_time):
    fig, ax = plt.subplots(figsize=(14, 6))
    x = df["t_stamp_dt"]
    noon_naive = noon.replace(tzinfo=None)
    
    if "Median_GHI" in df.columns:
        ax.plot(x, df["Median_GHI"], label="Median GHI", color="tab:blue", linestyle="-", linewidth=2)
    if "Median_POA" in df.columns:
        ax.plot(x, df["Median_POA"], label="Median POA", color="tab:orange", linestyle="-", linewidth=2)
        
    noon_str = pd.to_datetime(noon).round("1min").strftime("%H:%M")
    ax.axvline(noon_naive, color="red", linestyle="--", linewidth=2, label=f"Calc Solar Noon ({noon_str})")
    
    if "Median_GHI" in df.columns and not df["Median_GHI"].isna().all():
        max_ghi_idx = df["Median_GHI"].idxmax()
        max_ghi_time = df.loc[max_ghi_idx, "t_stamp_dt"]
        ax.axvline(max_ghi_time, color="tab:blue", linestyle=":", linewidth=2, label=f"Est GHI Noon ({max_ghi_time.strftime('%H:%M')})")

    if "Median_POA" in df.columns and not df["Median_POA"].isna().all():
        max_poa_idx = df["Median_POA"].idxmax()
        max_poa_time = df.loc[max_poa_idx, "t_stamp_dt"]
        ax.axvline(max_poa_time, color="tab:orange", linestyle=":", linewidth=2, label=f"Est POA Noon ({max_poa_time.strftime('%H:%M')})")

    format_time_axis(ax)
    ax.set_title(f"Median GHI & POA Overlay • {day}", fontsize=14)
    ax.set_ylabel("Irradiance [W/m²]", fontsize=12)
    ax.set_xlabel("Time (HH:MM)", fontsize=12)
    
    start_ts = pd.to_datetime(f"{day} {start_time}")
    end_ts = pd.to_datetime(f"{day} {end_time}")
    ax.set_xlim(start_ts, end_ts)
    ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0., fontsize=10)
    
    plt.tight_layout()
    save_path = output_folder / f"MEDIAN_GHI_POA_OVERLAY_{day}.png"
    plt.savefig(save_path, dpi=150)
    plt.close(fig)
    print(f"Saved: {save_path.name}")


# ------------------------------------------------------------
# 4) EXCEL EXPORT FUNCTIONS
# ------------------------------------------------------------

def write_excel_sheet(wb, sheet_name, date_val, records):
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Target Date:", str(date_val)])
    ws.append([])
    ws.append(["Type", "Sensor", "Solar Noon Est (EST)", "Irradiance (W/m²)", "Tilt (°)"])
    
    for r in records:
        ws.append([r["type"], r["sensor"], r["noon_est"], r["irr_val"], r["tilt_val"]])
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

def write_median_timeseries_sheet(wb, sheet_name, df, noon):
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Calculated Solar Noon:", pd.to_datetime(noon).round("1min").strftime("%H:%M:%S")])
    ws.append([])
    
    headers = ["t_stamp"]
    if "Median_GHI" in df.columns: headers.append("Median_GHI")
    if "Median_POA" in df.columns: headers.append("Median_POA")
    ws.append(headers)
    
    for _, row in df.iterrows():
        out_row = [row["t_stamp"]]
        if "Median_GHI" in df.columns: out_row.append(row["Median_GHI"])
        if "Median_POA" in df.columns: out_row.append(row["Median_POA"])
        ws.append(out_row)
        
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

def create_full_excel_report(output_path, records1, day1, records2, day2, df_median, noon2):
    try:
        from openpyxl import Workbook
        wb = Workbook()
        write_excel_sheet(wb, f"Peaks_{day1}", day1, records1)
        write_excel_sheet(wb, f"Peaks_{day2}", day2, records2)
        write_median_timeseries_sheet(wb, f"Median_Timeseries_{day2}", df_median, noon2)
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        wb.save(output_path)
        print(f"Exported Excel Report to: {output_path.name}")
    except ImportError:
        print("openpyxl not installed. Skipping Excel export.")
    except Exception as e:
        print(f"Failed to export Excel: {e}")


# ------------------------------------------------------------
# 5) MAIN EXECUTION
# ------------------------------------------------------------

def main():
    excel_files = list(input_folder.glob("*.xlsx"))
    if not excel_files:
        print(f"No files in {input_folder}"); sys.exit()

    print(f"Loading {len(excel_files)} files...")
    loaded = []
    for f in excel_files:
        try:
            df = prep_dataframe(pd.read_excel(f))
            # Updated to filter by SENSORS_TO_INCLUDE
            poa_cols = [c for c in df.columns if "/POA_" in c and "TILT" not in c and c in SENSORS_TO_INCLUDE]
            ghi_cols = [c for c in df.columns if "/GHI" in c and "TILT" not in c and c in SENSORS_TO_INCLUDE] 
            poa_tilt = [c for c in df.columns if "/POA_" in c and "TILT" in c and c.replace("_TILT_ANGLE","") in SENSORS_TO_INCLUDE]
            ghi_tilt = [c for c in df.columns if "/GHI_TILT" in c and c.replace("_TILT_ANGLE","") in SENSORS_TO_INCLUDE]
            
            # Explicit inclusion of RPOA tracking parameters 
            rpoa_cols = [c for c in df.columns if "/RPOA_" in c and "TILT" not in c and c in SENSORS_TO_INCLUDE]
            rpoa_tilt = [c for c in df.columns if "/RPOA_" in c and "TILT" in c and c.replace("_TILT_ANGLE","") in SENSORS_TO_INCLUDE]
            
            loaded.append({
                "file": f, "df": df,
                "poa_cols": poa_cols, "ghi_cols": ghi_cols,
                "poa_tilt": poa_tilt, "ghi_tilt": ghi_tilt,
                "rpoa_cols": rpoa_cols, "rpoa_tilt": rpoa_tilt
            })
        except Exception as e:
            print(f"Skip {f.name}: {e}")

    day1, day2 = parse_date(CompareDates[0]), parse_date(CompareDates[1])
    src1 = find_best_source_for_day(loaded, day1)
    src2 = find_best_source_for_day(loaded, day2)

    if not src1 or not src2:
        print("Missing data for comparison dates."); sys.exit()

    noon1 = get_precise_solar_noon(day1)
    noon2 = get_precise_solar_noon(day2)
    noon1_naive = noon1.replace(tzinfo=None)
    noon2_naive = noon2.replace(tzinfo=None)
    
    df1, df2 = src1["df_filtered"], src2["df_filtered"]
    row1 = df1.iloc[(df1["t_stamp_dt"] - noon1_naive).abs().argsort()[:1]]
    row2 = df2.iloc[(df2["t_stamp_dt"] - noon2_naive).abs().argsort()[:1]]

    print("\nGenerating Plots...")
    
    # Global GHI Comparison
    if src1["ghi_cols"] and src2["ghi_cols"]:
        plot_compare_2x2(
            df1, df2, src1["ghi_cols"], src1["ghi_tilt"], src2["ghi_cols"], src2["ghi_tilt"],
            day1, day2, noon1, noon2, row1, row2,
            title_prefix="GHI & Tilt Comparison", filename_suffix="Combined", mode="GHI"
        )
    
    # Global POA Comparison
    if src1["poa_cols"] and src2["poa_cols"]:
        plot_compare_2x2(
            df1, df2, src1["poa_cols"], src1["poa_tilt"], src2["poa_cols"], src2["poa_tilt"],
            day1, day2, noon1, noon2, row1, row2,
            title_prefix="POA & Tilt Comparison", filename_suffix="Combined", mode="POA"
        )
    
    # Global RPOA Comparison
    if src1["rpoa_cols"] and src2["rpoa_cols"]:
        plot_compare_2x2(
            df1, df2, src1["rpoa_cols"], src1["rpoa_tilt"], src2["rpoa_cols"], src2["rpoa_tilt"],
            day1, day2, noon1, noon2, row1, row2,
            title_prefix="RPOA & Tilt Comparison", filename_suffix="Combined", mode="RPOA"
        )
    
    # Station Subfolder Localized Comparisons
    stations = sorted(list(set([c.split("/")[0] for c in src1["poa_cols"]])))
    
    for st in stations:
        st_dir = output_folder / st
        st_dir.mkdir(exist_ok=True, parents=True)
        
        # 1. Localized POA Comparison
        p1 = [c for c in src1["poa_cols"] if c.startswith(st)]
        t1 = [c for c in src1["poa_tilt"] if c.startswith(st)]
        p2 = [c for c in src2["poa_cols"] if c.startswith(st)]
        t2 = [c for c in src2["poa_tilt"] if c.startswith(st)]
        
        if p1 and p2:
            fname_poa = f"POA_COMPARE_{day1}_VS_{day2}_{st}.png"
            source_path_poa = output_folder / fname_poa
            target_path_poa = st_dir / fname_poa
            
            plot_compare_2x2(
                df1, df2, p1, t1, p2, t2,
                day1, day2, noon1, noon2, row1, row2,
                title_prefix=f"POA & Tilt Comparison ({st})", filename_suffix=st, mode="POA"
            )
            # Route to Subfolder
            if target_path_poa.exists(): target_path_poa.unlink()
            if source_path_poa.exists(): source_path_poa.rename(target_path_poa)
            print(f"Moved {fname_poa} to {st_dir}")

        # 2. Localized RPOA Comparison
        r1 = [c for c in src1["rpoa_cols"] if c.startswith(st)]
        rt1 = [c for c in src1["rpoa_tilt"] if c.startswith(st)]
        r2 = [c for c in src2["rpoa_cols"] if c.startswith(st)]
        rt2 = [c for c in src2["rpoa_tilt"] if c.startswith(st)]
        
        if r1 and r2:
            fname_rpoa = f"RPOA_COMPARE_{day1}_VS_{day2}_{st}.png"
            source_path_rpoa = output_folder / fname_rpoa
            target_path_rpoa = st_dir / fname_rpoa
            
            plot_compare_2x2(
                df1, df2, r1, rt1, r2, rt2,
                day1, day2, noon1, noon2, row1, row2,
                title_prefix=f"RPOA & Tilt Comparison ({st})", filename_suffix=st, mode="RPOA"
            )
            # Route to Subfolder
            if target_path_rpoa.exists(): target_path_rpoa.unlink()
            if source_path_rpoa.exists(): source_path_rpoa.rename(target_path_rpoa)
            print(f"Moved {fname_rpoa} to {st_dir}")

    print("\nCalculating Medians & Plotting Overlay...")
    df1_medians = calculate_medians(df1, src1["ghi_cols"], src1["poa_cols"])
    df2_medians = calculate_medians(df2, src2["ghi_cols"], src2["poa_cols"])
    
    plot_median_ghi_poa_overlay(df1_medians, day1, noon1, start_time_limit, end_time_limit)
    plot_median_ghi_poa_overlay(df2_medians, day2, noon2, start_time_limit, end_time_limit)

    print("\nExporting Excel...")
    def get_records(row, cols, m_type):
        recs = []
        for c in cols:
            if c in row.index:
                val = row[c].values[0] if not pd.isna(row[c].values[0]) else ""
                t_col = f"{c}_TILT_ANGLE"
                tilt_val = row[t_col].values[0] if t_col in row.index and not pd.isna(row[t_col].values[0]) else ""
                noon_est = row["t_stamp"].values[0] if "t_stamp" in row.index else ""
                recs.append({
                    "type": m_type, "sensor": c, "noon_est": noon_est,
                    "irr_val": val, "tilt_val": tilt_val
                })
        return recs

    def export_analysis(t1, t2, filename_suffix):
        if not t1 or not t2: return
        r1 = df1.iloc[(df1["t_stamp_dt"] - t1).abs().argsort()[:1]]
        r2 = df2.iloc[(df2["t_stamp_dt"] - t2).abs().argsort()[:1]]
        
        recs1 = get_records(r1, src1["poa_cols"], "POA") + get_records(r1, src1["ghi_cols"], "GHI")
        recs2 = get_records(r2, src2["poa_cols"], "POA") + get_records(r2, src2["ghi_cols"], "GHI")
            
        fname = f"Solar_Noon_Analysis_{filename_suffix}.xlsx" if filename_suffix else "Solar_Noon_Analysis.xlsx"
        create_full_excel_report(output_folder / fname, recs1, day1, recs2, day2, df2_medians, noon2)

    export_analysis(noon1_naive, noon2_naive, "Calculated" if EXPORT_ESTIMATED_NOONS else "")

    if EXPORT_ESTIMATED_NOONS:
        t1_ghi = df1_medians.loc[df1_medians["Median_GHI"].idxmax(), "t_stamp_dt"] if "Median_GHI" in df1_medians.columns and not df1_medians["Median_GHI"].isna().all() else None
        t2_ghi = df2_medians.loc[df2_medians["Median_GHI"].idxmax(), "t_stamp_dt"] if "Median_GHI" in df2_medians.columns and not df2_medians["Median_GHI"].isna().all() else None
        export_analysis(t1_ghi, t2_ghi, "Est_GHI")
            
        t1_poa = df1_medians.loc[df1_medians["Median_POA"].idxmax(), "t_stamp_dt"] if "Median_POA" in df1_medians.columns and not df1_medians["Median_POA"].isna().all() else None
        t2_poa = df2_medians.loc[df2_medians["Median_POA"].idxmax(), "t_stamp_dt"] if "Median_POA" in df2_medians.columns and not df2_medians["Median_POA"].isna().all() else None
        export_analysis(t1_poa, t2_poa, "Est_POA")

    print("\nProcessing Complete.")

if __name__ == "__main__":
    main()